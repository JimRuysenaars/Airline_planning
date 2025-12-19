'''
Exercise 1B - Network Planning for Southern Star Airlines

Authors: 
Jim Rusenaars   (5309980)
Thijmen God     (5265762)
Lynn Vorgers    (5089301)

Date: November - December 2025

'''


# Import necessary libraries
import  pandas as pd
import  itertools
import  gurobipy as gp
import  numpy as np
from    matplotlib.patches import FancyArrowPatch
import  geopandas as gpd
from    matplotlib.lines import Line2D
import  openpyxl
from    openpyxl.styles import Font, PatternFill
import matplotlib.pyplot as plt


def get_ICAO(path):
    '''
    This function imports the ICAO codes of the airports.

    -Input-
    path :  path to airport information excel file

    -Output-
    icao:   list of ICAO codes
    '''

    df =    pd.read_excel(path, header=None)
    icao =  df.iloc[1, 1:].tolist()

    return icao

def get_demand(path):
    '''
    This function creates the demand matrix.

    -Input-
    path:       path to demand matrix excel file

    -Output- 
    demand_df:  demand matrix as a pandas DataFrame
    '''

    df =    pd.read_excel(path, header=None)    # Create dataframe from excel
    icao = df.iloc[0, 1:].tolist()              # ICAO codes

    demand_df = df.iloc[1:, 1:]                 # Create demand dataframe
    demand_df.index = icao                      # Set row labels
    demand_df.columns = icao                    # Set column labels

    demand_df = demand_df.astype(float)         # Ensure demand is in float form

    return demand_df

def create_AC(path):
    '''
    This function creates a dataframe with aircraft data.

    -Input-
    path:   path to aircraft data
    
    -Output-
    df:     Dataframe with aircraft data
    '''


    df = pd.read_excel(path)

    # First column is parameter names
    df = df.set_index(df.columns[0]).T

    # Remove any "Unit" column
    if "Unit" in df.columns:
        df = df.drop(columns=["Unit"])

    # Remove any index row called "Unit"
    df = df[~df.index.str.contains("Unit", case=False)]

    # Convert everything else to numeric
    df = df.apply(pd.to_numeric, errors='coerce')

    return df

def build_distance_matrix(path):
    '''
    This function calculates the great circle distance for each OD-pair

    -Input-
    path:       Path to airport data

    -Output-
    D:          Dataframe with distances between each OD-pair
    '''

    # Write Excel data to dataframe
    df = pd.read_excel(path, header=None)

    # Exctract data from dataframe
    icao = df.iloc[1, 1:].tolist()
    lat  = df.iloc[2, 1:].to_numpy(dtype=float)
    lon  = df.iloc[3, 1:].to_numpy(dtype=float)
    
    # Initialize matrix (empty)
    D = pd.DataFrame(index=icao, columns=icao, dtype=float)

    # Iterate over all i, j combinations
    for i, j in itertools.product(range(len(icao)), repeat=2):

        lat_i = lat[i]
        lon_i = lon[i]
        lat_j = lat[j]
        lon_j = lon[j]

        # Formula for great circle distance (given in assignment description)
        d_ij = 2 * np.arcsin(np.sqrt(np.sin(np.radians((lat_i - lat_j) / 2))**2 +
                            np.cos(np.radians(lat_i)) * np.cos(np.radians(lat_j)) * np.sin(np.radians((lon_i - lon_j) / 2))**2))

        d_ij = d_ij * 6371  # Radius of Earth in kilometers
        D.iloc[i, j] = d_ij

    return D

def ICAO_coordinates(path):
    '''
    This function creates a dataframe with the coordinates of each airport.

    -Input-
    path:   path to airport data

    -Output-
    coordinates:   Dataframe with latitude and longitude of each airport
    '''

    # Extract Excel data to dataframe
    df = pd.read_excel(path, header=None)

    # Extract data from dataframe
    icao = df.iloc[1, 1:].tolist()
    lat  = df.iloc[2, 1:].to_numpy(dtype=float)
    lon  = df.iloc[3, 1:].to_numpy(dtype=float)

    coordinates = pd.DataFrame(index=icao, columns=['latitude', 'longitude'], dtype=float)
    for i in range(len(icao)):
        coordinates.loc[icao[i], 'latitude'] = lat[i]
        coordinates.loc[icao[i], 'longitude'] = lon[i]

    return coordinates



def create_OD(path_distance, path_demand):
    '''
    This function creates the large OD dataframe with distance, demand and yield for each OD-pair.

    -Input-
    path_distance:  path to airport data
    path_demand:    path to demand data

    -Output-
    OD:             Dataframe with distance, demand and yield for each OD-pair

    '''

    Distance_matrix =   build_distance_matrix(path_distance)
    Demand_matrix  =    get_demand(path_demand)
    I =                 get_ICAO(path_distance)
    J =                 get_ICAO(path_distance)

    rows = []
    for i, j in itertools.product(I, J):
        if i == j:
            # Skip same airport pairs
            continue  

        else:
            distance = Distance_matrix.loc[i, j]  
            demand   = Demand_matrix.loc[i, j]
            yield_   = 5.9 * distance ** -0.76 + 0.043 if distance > 0 else 0  # Avoid division by zero

            # Append data to list
            rows.append({
                "i": i,
                "j": j,
                "distance": distance,
                "demand": demand,
                "yield":  yield_
            })

    OD = pd.DataFrame(rows).set_index(["i", "j"]).sort_index()

    return OD

def create_AP(path, hub='LEMF'):
    '''
    This function creates a dataframe with airport data, including runway length and available slots.

    -Input-
    path:   path to airport data

    -Output-
    AP:     Dataframe with airport data
    '''

    # Extract Excel data to variables
    df =                    pd.read_excel(path, header=None)
    icao =                  df.iloc[1, 1:].tolist()
    runway_length_raw =     df.iloc[4, 1:]
    runway_length =         pd.to_numeric(runway_length_raw, errors='coerce').fillna(0).to_numpy()
    available_slots_raw =   df.iloc[5, 1:]
    available_slots =       pd.to_numeric(available_slots_raw, errors='coerce').fillna(0).to_numpy()
    AP =                    pd.DataFrame(index=icao, columns=['runway_length', 'available_slots'], dtype=float)

    # Loop over all airports and fill dataframe
    for i in range(len(icao)):

        # Runway length
        AP.loc[icao[i], 'runway_length'] = runway_length[i]

        # Hub indicator and available slots
        if icao[i] == hub:
            AP.loc[icao[i], 'G'] = 0

            # Hub airport has infinite slots
            AP.loc[icao[i], 'available_slots'] = 9999

        else:
            AP.loc[icao[i], 'G'] = 1
            AP.loc[icao[i], 'available_slots'] = available_slots[i]   

    return AP

def build_a_ijk(OD, AC):
    '''
    This function builds the a_ijk parameter dictionary. This parameter indicates whether an aircraft type k can serve the OD-pair (i,j) based on its range.

    -Input-
    OD:     Dataframe with OD data
    AC:     Dataframe with aircraft data

    -Output-
    a_ijk:  dictionary with a_ijk values
    '''
    # Extract I, J, K sets
    I = OD.index.get_level_values(0).unique().to_list()                         
    J = OD.index.get_level_values(1).unique().to_list()                         
    K = AC.index  

    # Build a_ijk dictionary                                                              
    a_ijk = {}
    for i in I:
        for j in J:
            for k in K:
                if i == j:
                    continue
                else:
                    if OD.loc[(i, j), "distance"] <= AC.loc[k, "maximum_range"]:

                        # Aircraft can serve the route
                        a_ijk[(i, j, k)] = 10000 # Serves as a big M
                    else:
                        # Aircraft cannot serve the route
                        a_ijk[(i, j, k)] = 0

    return a_ijk

# Build dataframes
AC = create_AC('Problem 1 - Data\\AircraftData.xlsx')
OD = create_OD('Problem 1 - Data\\airport_data.xlsx', 'Problem 1 - Data\\demand_per_week.xlsx')
AP = create_AP('Problem 1 - Data\\airport_data.xlsx')
a_ijk = build_a_ijk(OD, AC)


# Initialize gurobipy model
model = gp.Model("Airline_Planning")

# Parameters
fuel_cost =         1.42        # EUR/gallon
utilization_time =  10 * 7      # 10 hours of operstions per day, 7 days a week -> 70 hours per week
LF =                0.75        # Load factor
eps = 1e-6                      # Machine epsilon to compare values close to zero

# Sets
pairs = OD.index.tolist()       # List of (i,j) pairs
K =     AC.index                # Set of aircraft types k 

# Decision variables
x = model.addVars(pairs, name="xij", vtype=gp.GRB.INTEGER, lb=0)            # Direct flow from airport i to airport j
w = model.addVars(pairs, name="wij", vtype=gp.GRB.INTEGER, lb=0)            # Flow from airport i to airport j that transfers at the hub
y = model.addVars(K, name="y", vtype=gp.GRB.INTEGER, lb=0)                  # Number of aircraft of type k
z = model.addVars( [(i, j, k)                                               # Number of flights from airport i to airport j with aircraft type k
                    for (i, j) in pairs for k in K], name="zijk", 
                    vtype=gp.GRB.INTEGER, lb=0)

# Revenue
revenue = gp.quicksum(
    (5.9 * (OD.loc[p, "distance"] + eps) ** (-0.76) + 0.043) * (x[p] + w[p]) * OD.loc[p, "distance"]
    for p in pairs
)

# Costs for operations and leasing
costs = (gp.quicksum(
    z[i, j, k] * (                                                                                      # Number of flights on (i,j) with aircraft k
        AC.loc[k, "fixed_operating_cost"]                                                               # Fixed operating cost per flight                                     
        + AC.loc[k, "time_cost_parameter"] * (OD.loc[(i, j), "distance"] / AC.loc[k, "speed"])          # Time cost per flight   
        + AC.loc[k, "fuel_cost_parameter"] * fuel_cost * OD.loc[(i, j), "distance"] / 1.585             # Fuel cost per flight
    ) * 0.7                                                                                             # 30% cost reduction due to economies of scale                    
    for (i, j) in pairs
    for k in K
) + 
gp.quicksum(y[k] * AC.loc[k, "weekly_lease_cost"] for k in K))                                          # Leasing cost per week

# Set objective function
model.setObjective(revenue - costs, gp.GRB.MAXIMIZE)

 # Set MIP gap
model.setParam('MIPGap', 0.004)                                                                        

# --- Constraints ----

# C1: Number of transported passengers lower than demand
model.addConstrs(
    ( x[p] + w[p] <= OD.loc[p, "demand"] for p in pairs ),
    name="C1_Demand"
)

# C1*: W is not larger than demand and is 0 when i or j is hub
model.addConstrs(
    ( w[p] <= OD.loc[p, "demand"] * AP.loc[p[1], "G"] * AP.loc[p[0], "G"] for p in pairs ),
    name="C1*_Transfer_via_hub"
)

# C2: Capacity per flight
for (i, j) in pairs:
    model.addConstr(
        x[(i, j)] + 
            gp.quicksum( w[(i, m)] * (1 - AP.loc[j, "G"]) 
                        for m in AP.index if m != j if m != i)
             + gp.quicksum( w[(m, j)] * ( 1- AP.loc[i, "G"]) 
                           for m in AP.index if m != j if m != i)
        <= gp.quicksum(z[i, j, k] * AC.loc[k, "seats"] * LF 
                       for k in K),
        name=f"C2_Capacity_{i}_{j}"
    )

# C3: Every flight goes back and forth
for i in AP.index:
    for k in K:
        model.addConstr(
                gp.quicksum(z[i, j, k] 
                            for j in AP.index if j != i)
                ==
                gp.quicksum(z[j, i, k] 
                            for j in AP.index if j != i),
                name=f"C3_RoundTrip_{i}_{k}"
            )

# C4: Fleet availability: Operating time is smaller than available time
for k in K:
    model.addConstr(
        gp.quicksum(
            z[i, j, k] * (
                OD.loc[(i, j), "distance"] / AC.loc[k, "speed"] +
                AC.loc[k, "average_tat"] / 60 * (1 + 0.5 * (1 - AP.loc[j, "G"]))
            )
            for (i, j) in pairs
        )
        <= utilization_time * y[k],
        name=f"C4_Fleet_{k}"
    )

# C5: Range constraint
for (i, j) in pairs:
    for k in K:
        model.addConstr(
            z[i, j, k] <= a_ijk[(i, j, k)],
            name=f"C5_Range_{i}_{j}_{k}"
        )

# C6: All flights start or end at hub
for (i, j) in pairs:
    for k in K:
        model.addConstr(
            z[i, j, k] * AP.loc[i, "G"] * AP.loc[j, "G"] <= eps,
            name=f"C6_Hub_{i}_{j}_{k}"
        )

# C7 : Runway length constraint
for (i, j) in pairs:
    for k in K:
        model.addConstr(
            z[i, j, k] * AC.loc[k, "runway_required"] <= AP.loc[j, "runway_length"] *  z[i, j, k]  ,
            name=f"C7_Runway_{i}_{j}_{k}"
        )

# C8: Available slots for landing constraint
for j in AP.index:
    model.addConstr(
        gp.quicksum(z[i, j, k] for i in AP.index if i != j for k in K) <= AP.loc[j, "available_slots"],
        name=f"C8_Available_slots_{j}")


# Optimize model
model.optimize()

# ==== Process results ====

# Retrieve and print results
if model.status == gp.GRB.OPTIMAL:
    print(f"Optimal objective value: {model.objVal}")
    for v in model.getVars():
        if v.X > eps:
            print(f"{v.VarName}: {v.X}")

# Write model to file
model.write("model.lp")

z_values = {}
for z in model.getVars():
    if not z.VarName.startswith("z"):
        continue

    if abs(z.X) < eps:
        continue

    for constr in model.getConstrs():
        coef = model.getCoeff(constr, z)
        z_values[z.VarName] = z.X


def make_map_AC_type(z_values):
    '''
    This function colors an aircraft map with the network according to the aircraft type used.

    -Input-
    z_values:   dictionary with z variable names and their values

    -Output-
    A map plot showing routes colored by aircraft type
    '''


    coordinates_df = ICAO_coordinates('Problem 1 - Data\\airport_data.xlsx')
    world = gpd.read_file('https://naciscdn.org/naturalearth/110m/cultural/ne_110m_admin_0_countries.zip')

    fig, ax = plt.subplots(figsize=(14, 10))

    # Plot all airports
    ax.scatter(coordinates_df['longitude'], coordinates_df['latitude'], 
               color='black', s=30, zorder=5, label='Airports')


        
    oceans = gpd.read_file(
    "https://naciscdn.org/naturalearth/110m/physical/ne_110m_ocean.zip"
).to_crs("EPSG:4326")

    # Load world map from Natural Earth data
    # --- WATER ---
    oceans.plot(
        ax=ax,
        color='lightblue',
        edgecolor='none',
        zorder=0
    )

    # --- LAND / COUNTRIES ---
    world.plot(
        ax=ax,
        column= None,
        color='#f7c97f',
        edgecolor='black',
        linewidth=0.5,
        legend=True,
        zorder=1
    )


    # Plot flight routes where at least one flight is operated
    for z_var_name, z_value in z_values.items():
        if z_value > 0.8:
            parts = z_var_name.split('[')[1].rstrip(']').split(',')
            airport_i = parts[0].strip()
            airport_j = parts[1].strip()
            
            if airport_i in coordinates_df.index and airport_j in coordinates_df.index:
                lon_i = coordinates_df.loc[airport_i, 'longitude']
                lat_i = coordinates_df.loc[airport_i, 'latitude']
                lon_j = coordinates_df.loc[airport_j, 'longitude']
                lat_j = coordinates_df.loc[airport_j, 'latitude']
                
                # Extract aircraft type from z variable name
                parts = z_var_name.split('[')[1].rstrip(']').split(',')
                aircraft_type = parts[2].strip()
                
                # Define color map for aircraft types
                color_map = {
                    'AC1': 'navy',
                    'AC2': 'darkgreen',
                    'AC3': 'purple',
                    'AC4': 'red'
                }
                
                line_color = color_map.get(aircraft_type, 'gray')
                
                arrow = FancyArrowPatch((lon_i, lat_i), (lon_j, lat_j),
                                        arrowstyle='-', mutation_scale=20, 
                                        color=line_color, alpha=0.6, linewidth=3, zorder=4)
                ax.add_patch(arrow)

    # Add airport labels
    for airport, row in coordinates_df.iterrows():
        ax.annotate(airport, (row['longitude'], row['latitude']), 
                    xytext=(5, 5), textcoords='offset points', fontsize=8)

        # Create legend for aircraft types
        legend_elements = [Line2D([0], [0], color='navy', lw=3, label='AC1'),
                           Line2D([0], [0], color='darkgreen', lw=3, label='AC2'),
                           Line2D([0], [0], color='purple', lw=3, label='AC3'),
                           Line2D([0], [0], color='red', lw=3, label='AC4')]
        ax.legend(handles=legend_elements, loc='upper left')

    ax.set_xlabel('Longitude')
    ax.set_ylabel('Latitude')
    ax.set_title('Southern Star Airlines aircraft type map')
    ax.set_xlim(-30, 35)
    ax.axis('off')
    ax.set_ylim(30, 70)

    # Plot the world map as background
    world.boundary.plot(ax=ax, linewidth=1, color='black')

    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.show()

def write_schedule(z_values, filename="flight_schedule.xlsx"):
    '''
    This function writes the flight schedule to an Excel file.
    
    -Input-
    z_values:   dictionary with z variable names and their values
    filename:   name of the output Excel file

    -Output-
    An Excel file containing the flight schedule
    '''

    schedule_data = []
    for z_var_name, z_value in z_values.items():
        parts = z_var_name.split('[')[1].rstrip(']').split(',')
        airport_i = parts[0].strip()
        airport_j = parts[1].strip()
        aircraft_type = parts[2].strip()
        
        schedule_data.append({
            'Origin': airport_i,
            'Destination': airport_j,
            'Aircraft': aircraft_type,
            'Flights': z_value
        })
    
    # Create DataFrame and write to Excel
    df = pd.DataFrame(schedule_data)
    df.to_excel(filename, index=False, sheet_name='Schedule')
    
    # Format Excel file
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    wb.save(filename)

def make_map_frequency(z_values):
    '''
    This function colors an aircraft map with the network according to the flight frequency.

    -Input-
    z_values:   dictionary with z variable names and their values

    -Output-
    A map plot showing routes colored by flight frequency
    '''


    linewidth = 3
    coordinates_df = ICAO_coordinates('Problem 1 - Data\\airport_data.xlsx')
    world = gpd.read_file('https://naciscdn.org/naturalearth/110m/cultural/ne_110m_admin_0_countries.zip')

    fig, ax = plt.subplots(figsize=(14, 10))

    # Plot all airports
    ax.scatter(coordinates_df['longitude'], coordinates_df['latitude'], 
               color='black', s=30, zorder=5, label='Airports')

    oceans = gpd.read_file(
    "https://naciscdn.org/naturalearth/110m/physical/ne_110m_ocean.zip"
).to_crs("EPSG:4326")

    # Load world map from Natural Earth data
    # --- WATER ---
    oceans.plot(
        ax=ax,
        color='lightblue',
        edgecolor='none',
        zorder=0
    )

    # --- LAND / COUNTRIES ---
    world.plot(
        ax=ax,
        column= None,
        color='#f7c97f',
        edgecolor='black',
        linewidth=0.5,
        legend=True,
        zorder=1
    )

    # Plot routes where at least one flight is operated
    for z_var_name, z_value in z_values.items():
        if z_value > 0.8:
            parts = z_var_name.split('[')[1].rstrip(']').split(',')
            airport_i = parts[0].strip()
            airport_j = parts[1].strip()
            
            if airport_i in coordinates_df.index and airport_j in coordinates_df.index:
                lon_i = coordinates_df.loc[airport_i, 'longitude']
                lat_i = coordinates_df.loc[airport_i, 'latitude']
                lon_j = coordinates_df.loc[airport_j, 'longitude']
                lat_j = coordinates_df.loc[airport_j, 'latitude']
                color_map = {'AC1': 'midnightblue', 'AC2': 'darkgreen', 'AC3': 'purple', 'AC4': 'red'}
                # Color based on frequency (z_value)
                if z_value <= 1.5:
                    line_color = 'midnightblue'
                    linewidth = linewidth
                elif z_value <= 2.5:
                    line_color = 'darkgreen'
                    linewidth = linewidth
                elif z_value <= 3.5:
                    line_color = 'purple'
                    linewidth = linewidth
                else:
                    line_color = 'red'
                    linewidth = linewidth
                
                arrow = FancyArrowPatch((lon_i, lat_i), (lon_j, lat_j),
                                        arrowstyle='-', mutation_scale=20, 
                                        color=line_color, alpha=0.6, linewidth=linewidth, zorder=4)
                ax.add_patch(arrow)

    # Add airport labels
    for airport, row in coordinates_df.iterrows():
        ax.annotate(airport, (row['longitude'], row['latitude']), 
                    xytext=(5, 5), textcoords='offset points', fontsize=8)

    # Create legend for frequencies
    legend_elements = [Line2D([0], [0], color='midnightblue', lw=linewidth, label='1 time/week'),
                       Line2D([0], [0], color='darkgreen', lw=linewidth, label='2 times/week'),
                       Line2D([0], [0], color='purple', lw=linewidth, label='3 times/week'),
                       Line2D([0], [0], color='red', lw=linewidth, label='4+ times/week')]
    
    ax.legend(handles=legend_elements, loc='upper left')
    ax.set_xlabel('Longitude')
    ax.set_ylabel('Latitude')
    ax.set_title('Southern Star Airlines flight frequency map')
    ax.axis('off')
    ax.set_xlim(-30, 35)
    ax.set_ylim(30, 70)
    world.boundary.plot(ax=ax, linewidth=1, color='black')
    ax.grid(True, alpha=0.3)

    plt.tight_layout()
    plt.show()

def plot_range_map():
    '''
    This function plots the range circles for each aircraft type from the hub airport LEMF.

    -Input-

    -Output-    
    A map plot showing range circles for each aircraft type from LEMF
    '''

    coordinates_df = ICAO_coordinates('Problem 1 - Data\\airport_data.xlsx')
    world = gpd.read_file('https://naciscdn.org/naturalearth/110m/cultural/ne_110m_admin_0_countries.zip')

    fig, ax = plt.subplots(figsize=(14, 10))

    # Plot all airports
    ax.scatter(coordinates_df['longitude'], coordinates_df['latitude'], 
               color='black', s=30, zorder=5, label='Airports')

    oceans = gpd.read_file(
    "https://naciscdn.org/naturalearth/110m/physical/ne_110m_ocean.zip"
    ).to_crs("EPSG:4326")

    # --- WATER ---
    oceans.plot(
        ax=ax,
        color='lightblue',
        edgecolor='none',
        zorder=0
    )

    # --- LAND / COUNTRIES ---
    world.plot(
        ax=ax,
        column=None,
        color='#f7c97f',
        edgecolor='black',
        linewidth=0.5,
        legend=True,
        zorder=1
    )

    # Get LEMF coordinates
    lemf_lon = coordinates_df.loc['LEMF', 'longitude']
    lemf_lat = coordinates_df.loc['LEMF', 'latitude']

    # Plot range circles for each aircraft
    color_map = {'AC1': 'navy', 'AC2': 'darkgreen', 'AC3': 'purple', 'AC4': 'red'}

    for aircraft in AC.index:
        max_range_km = AC.loc[aircraft, 'maximum_range']
        # Convert km to degrees (approximate: 1 degree ≈ 111 km)
        max_range_deg = max_range_km / 111
        

        circle = plt.Circle((lemf_lon, lemf_lat), max_range_deg, 
                           color=color_map.get(aircraft, 'gray'), 
                           fill=False, linewidth=2, alpha=0.6, zorder=3,
                           label=f'{aircraft} (range: {max_range_km:.0f}km)')
        ax.add_patch(circle)

    # Add airport labels
    for airport, row in coordinates_df.iterrows():
        ax.annotate(airport, (row['longitude'], row['latitude']), 
                    xytext=(5, 5), textcoords='offset points', fontsize=8)

    ax.set_xlabel('Longitude')
    ax.set_ylabel('Latitude')
    ax.set_title('Aircraft Range from LEMF')
    ax.set_xlim(-30, 35)
    ax.set_ylim(30, 70)

    world.boundary.plot(ax=ax, linewidth=1, color='black')
    ax.legend(loc='upper left')
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.show()

def write_slack_to_excel():
    '''
    This function writes the constraint slacks to an Excel file.

    -Input-

    -Output-
    An Excel file containing the constraint slacks
    '''

    slack_data = []
    for constr in model.getConstrs():
        constraint_type = constr.ConstrName[:2]
        slack_data.append({
            'Type': constraint_type,
            'Constraint Name': constr.ConstrName,
            'Slack': constr.Slack
        })

    slack_df = pd.DataFrame(slack_data)
    slack_df.to_excel("constr_slacks.xlsx", index=False, sheet_name='Constraints')

    # Format Excel file
    wb = openpyxl.load_workbook("constr_slacks.xlsx")
    ws = wb.active
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15

    # Add header formatting
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    wb.save("constr_slacks.xlsx")
